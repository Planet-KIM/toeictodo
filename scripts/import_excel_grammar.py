import sqlite3
import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, 'TOEIC_850_접속사_접속부사_전치사_관계사_완전정리_예문포함.xlsx')
DB_PATH = os.path.join(BASE_DIR, 'toeic.db')

SHEET_CONFIGS = [
    {
        'sheet': '부사절_접속사',
        'default_pos': '접속사',
        'word_col': '표현',
        'meaning_col': '뜻',
        'topic_col': '의미 분류',
        'collocation_col': '기본 구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': '예문 해석',
        'trap_col': 'TOEIC 출제 포인트',
        'prio_col': '중요도'
    },
    {
        'sheet': '접속부사',
        'default_pos': '접속부사',
        'word_col': '표현',
        'meaning_col': '뜻',
        'topic_col': '의미 분류',
        'collocation_col': '기본 구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': '예문 해석',
        'trap_col': 'TOEIC 출제 포인트',
        'prio_col': '중요도'
    },
    {
        'sheet': '일반_접속사',
        'default_pos': '접속사',
        'word_col': '표현',
        'meaning_col': '뜻',
        'topic_col': '의미 분류',
        'collocation_col': '기본 구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': '예문 해석',
        'trap_col': 'TOEIC 출제 포인트',
        'prio_col': '중요도'
    },
    {
        'sheet': '전치사',
        'default_pos': '전치사',
        'word_col': '표현',
        'meaning_col': '뜻',
        'topic_col': '의미 분류',
        'collocation_col': '기본 구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': '예문 해석',
        'trap_col': 'TOEIC 출제 포인트',
        'prio_col': '중요도'
    },
    {
        'sheet': '명사절_접속사_850',
        'default_pos': '접속사',
        'word_col': '표현',
        'meaning_col': '뜻',
        'topic_col': '종류',
        'collocation_col': '기본 구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': '예문 해석',
        'trap_col': '출제 포인트',
        'prio_col': '중요도'
    },
    {
        'sheet': '관계사_850',
        'default_pos': '관계사',
        'word_col': '표현',
        'meaning_col': '뜻',
        'topic_col': '종류',
        'collocation_col': '기본 구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': '예문 해석',
        'trap_col': '출제 포인트',
        'prio_col': '중요도'
    },
    {
        'sheet': '다품사_함정_850',
        'default_pos': None, # read from '가능 품사'
        'word_col': '표현',
        'meaning_col': None,
        'topic_col': None,
        'collocation_col': '구조',
        'example_en_col': 'TOEIC 스타일 예문',
        'example_ko_col': None,
        'trap_col': '핵심 구분',
        'prio_col': '중요도',
        'pos_col': '가능 품사'
    }
]

def run_import():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    inserted_count = 0
    updated_count = 0

    for cfg in SHEET_CONFIGS:
        sheetname = cfg['sheet']
        if sheetname not in wb.sheetnames:
            print(f"Skipping missing sheet: {sheetname}")
            continue

        sheet = wb[sheetname]
        headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
        col_map = {str(h).strip(): idx + 1 for idx, h in enumerate(headers) if h}

        print(f"Processing sheet '{sheetname}'...")

        for r in range(2, sheet.max_row + 1):
            def get_val(col_name):
                if not col_name or col_name not in col_map:
                    return ""
                val = sheet.cell(r, col_map[col_name]).value
                return str(val).strip() if val is not None else ""

            word = get_val(cfg['word_col'])
            if not word:
                continue

            word_lower = word.lower()
            pos = get_val(cfg.get('pos_col')) if cfg.get('pos_col') else cfg['default_pos']
            meaning = get_val(cfg['meaning_col']) or "의미 정보"
            topic = get_val(cfg['topic_col']) or "접속사·전치사"
            collocation = get_val(cfg['collocation_col'])
            example_en = get_val(cfg['example_en_col'])
            example_ko = get_val(cfg['example_ko_col'])
            trap_point = get_val(cfg['trap_col'])
            prio_raw = get_val(cfg['prio_col']).upper()
            priority = prio_raw if prio_raw in ['A', 'B', 'C'] else 'A'

            # Check if word exists in DB
            cursor.execute("SELECT id, pos, meaning, collocation, trap_point FROM words WHERE LOWER(word) = ?", (word_lower,))
            row = cursor.fetchone()

            if row:
                word_id = row[0]
                existing_pos = row[1] or ""
                
                # Combine POS if different
                new_pos = existing_pos
                if pos and pos not in existing_pos:
                    new_pos = f"{existing_pos}, {pos}".strip(", ") if existing_pos else pos

                cursor.execute("""
                    UPDATE words 
                    SET pos = ?, 
                        collocation = CASE WHEN LENGTH(?) > 0 THEN ? ELSE collocation END,
                        trap_point = CASE WHEN LENGTH(?) > 0 THEN ? ELSE trap_point END,
                        example_en = CASE WHEN LENGTH(?) > 0 THEN ? ELSE example_en END,
                        example_ko = CASE WHEN LENGTH(?) > 0 THEN ? ELSE example_ko END,
                        priority = ?
                    WHERE id = ?
                """, (new_pos, collocation, collocation, trap_point, trap_point, example_en, example_en, example_ko, example_ko, priority, word_id))
                updated_count += 1
            else:
                cursor.execute("""
                    INSERT INTO words (word, pos, meaning, priority, topic, collocation, trap_point, example_en, example_ko)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (word, pos, meaning, priority, topic, collocation, trap_point, example_en, example_ko))
                inserted_count += 1

    conn.commit()
    conn.close()
    print(f"=== Import Completed! Successfully inserted {inserted_count} new words and updated {updated_count} existing words in DB! ===")

if __name__ == '__main__':
    run_import()
